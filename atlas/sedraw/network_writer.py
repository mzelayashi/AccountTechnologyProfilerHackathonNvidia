"""Write a network-topology intake .xlsx in the form layout NetworkReader expects.

Counterpart to network_reader.NetworkReader.read_network_topology — lets the in-GUI editor
persist edits back to the actual Excel file. Layout mirrors the canonical generator
(SEDraw Cache/create_network_test_data.py): column A = labels, column B = values.

  Site: <label>            | DC | Branch
  <Category name>          |
    Qty x Model:           | <value>
    Connected to:          | <value>
    End of Support Date:   | <value>
    Rating (1-5):          | <value>
    Status:                | <value>      (added so 'status' round-trips losslessly)
    Notes:                 | <value>
  (blank row between categories; blank row between sites)

Round-trips cleanly through NetworkReader (its FIELD_LABELS includes 'status:').
"""
from __future__ import annotations

from pathlib import Path

# (column-A label, category-dict key) — write order for each category's fields.
FIELD_ORDER = [
    ("Qty x Model:", "qty_model"),
    ("Connected to:", "connected_to"),
    ("End of Support Date:", "end_of_support"),
    ("Rating (1-5):", "rating"),
    ("Status:", "status"),
    ("Notes:", "notes"),
]


def _site_type_label(site_type: str) -> str:
    return "DC" if (site_type or "").lower() == "dc" else "Branch"


def write_network_topology(network_data: dict, out_path: Path) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Network Topology"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    site_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    fill_dc = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    fill_br = PatternFill(start_color="4E342E", end_color="4E342E", fill_type="solid")
    cat_font = Font(name="Calibri", size=11, bold=True, color="1E3A5F")
    cat_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
    field_font = Font(name="Calibri", size=10, color="555555")
    value_font = Font(name="Calibri", size=10)

    r = 1
    for site in (network_data or {}).get("sites", []):
        label = (site.get("label") or "Site").strip()
        st = _site_type_label(site.get("site_type"))
        fill = fill_dc if st == "DC" else fill_br
        a = ws.cell(row=r, column=1, value=f"Site: {label}"); a.font = site_font; a.fill = fill
        b = ws.cell(row=r, column=2, value=st); b.font = site_font; b.fill = fill
        r += 1

        for cat in site.get("categories", []):
            name = (cat.get("category") or "").strip()
            if not name:
                continue
            ch = ws.cell(row=r, column=1, value=name); ch.font = cat_font; ch.fill = cat_fill
            ws.cell(row=r, column=2).fill = cat_fill
            r += 1
            for col_label, key in FIELD_ORDER:
                ws.cell(row=r, column=1, value=col_label).font = field_font
                val = cat.get(key, "")
                if val is None:
                    val = ""
                ws.cell(row=r, column=2, value=str(val)).font = value_font
                r += 1
            r += 1   # blank separator between categories
        r += 1       # blank separator between sites

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
