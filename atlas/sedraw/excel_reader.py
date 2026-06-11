"""Excel input reader and validator for SE Draw."""
from pathlib import Path
from typing import List, Dict, Tuple


class ExcelReader:
    """Read and validate Vision Board Excel input files."""

    REQUIRED_SYSTEMS_COLS = ['system_name', 'vendor', 'role', 'status', 'tier']
    REQUIRED_FLOWS_COLS = ['source', 'target']
    VALID_STATUSES = {'current', 'retiring', 'new', 'planned'}
    VALID_TIERS = {'source', 'hub', 'consumer', 'integration', 'ai'}
    VALID_FLOW_TYPES = {'data_feed', 'etl', 'api', 'integration', 'sync', 'query'}
    VALID_LINE_STYLES = {'solid', 'dashed', 'dotted'}

    def read_vision_board(self, excel_path: Path) -> Tuple[List[Dict], List[Dict]]:
        """
        Read a vision board Excel file with Systems and Data Flows sheets.

        Returns:
            (systems_list, data_flows_list)

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required columns or sheets are missing
        """
        import openpyxl

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        # Read Systems sheet
        if 'Systems' not in wb.sheetnames:
            raise ValueError(f"Missing 'Systems' sheet in {excel_path.name}")
        systems = self._read_systems_sheet(wb['Systems'])

        # Read Data Flows sheet
        if 'Data Flows' not in wb.sheetnames:
            raise ValueError(f"Missing 'Data Flows' sheet in {excel_path.name}")
        flows = self._read_flows_sheet(wb['Data Flows'])

        wb.close()

        # Validate cross-references
        warnings = self._validate_references(systems, flows)
        for w in warnings:
            print(f"  Warning: {w}")

        return systems, flows

    def _read_systems_sheet(self, ws) -> List[Dict]:
        """Read the Systems sheet into a list of dicts."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Systems sheet is empty")

        # Normalize headers
        headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]

        # Check required columns
        for col in self.REQUIRED_SYSTEMS_COLS:
            if col not in headers:
                raise ValueError(f"Systems sheet missing required column: '{col}'")

        systems = []
        for row in rows[1:]:
            if not row or all(cell is None for cell in row):
                continue
            entry = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    entry[header] = str(row[i]).strip() if row[i] is not None else ''
            # Normalize status and tier
            entry['status'] = self._normalize_value(entry.get('status', ''), self.VALID_STATUSES, 'current')
            entry['tier'] = self._normalize_value(entry.get('tier', ''), self.VALID_TIERS, 'source')
            if entry.get('system_name'):
                systems.append(entry)

        return systems

    def _read_flows_sheet(self, ws) -> List[Dict]:
        """Read the Data Flows sheet into a list of dicts."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Data Flows sheet is empty")

        headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]

        for col in self.REQUIRED_FLOWS_COLS:
            if col not in headers:
                raise ValueError(f"Data Flows sheet missing required column: '{col}'")

        flows = []
        for row in rows[1:]:
            if not row or all(cell is None for cell in row):
                continue
            entry = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    entry[header] = str(row[i]).strip() if row[i] is not None else ''
            # Normalize optional fields
            entry['flow_type'] = self._normalize_value(
                entry.get('flow_type', ''), self.VALID_FLOW_TYPES, 'data_feed')
            entry['line_style'] = self._normalize_value(
                entry.get('line_style', ''), self.VALID_LINE_STYLES, 'solid')
            if entry.get('source') and entry.get('target'):
                flows.append(entry)

        return flows

    def _validate_references(self, systems: List[Dict], flows: List[Dict]) -> List[str]:
        """Check that flow source/target values match a system_name."""
        system_names = {s['system_name'].lower() for s in systems}
        warnings = []
        for flow in flows:
            src = flow.get('source', '').lower()
            tgt = flow.get('target', '').lower()
            if src and src not in system_names:
                warnings.append(f"Flow source '{flow['source']}' not found in Systems sheet")
            if tgt and tgt not in system_names:
                warnings.append(f"Flow target '{flow['target']}' not found in Systems sheet")
        return warnings

    def _normalize_value(self, value: str, valid_set: set, default: str) -> str:
        """Normalize a value to lowercase and validate against allowed set."""
        normalized = value.strip().lower().replace(' ', '_')
        if normalized in valid_set:
            return normalized
        if normalized:
            print(f"  Warning: Unknown value '{value}', defaulting to '{default}'")
        return default
